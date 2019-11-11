import openpyxl
from googletrans import Translator
translator=Translator()
orgFileLocation=("E:\\notes\\Python\\document translation tool\\file_example_XLSX_100.xlsx")
xlWorkbook=openpyxl.load_workbook(orgFileLocation)
xlWorkSheetLst=xlWorkbook.sheetnames
print(xlWorkSheetLst)
xlWorkBookMetaData=[]
for index in xlWorkSheetLst:
    xlWorkSheet=xlWorkbook[index]
    xlWorkBookMetaData.append([index,xlWorkSheet.max_row,xlWorkSheet.max_column])
print(xlWorkBookMetaData)
xlWorkSheetLst=[]
xlWorkBookData=[]
for index in xlWorkBookMetaData:
    xlWorkSheet=xlWorkbook[index[0]]
    for rowIndex in range(1,index[1]+1):
        for colIndex in range(1,index[2]+1):
            tempData=xlWorkSheet.cell(row=rowIndex,column=colIndex)
            tempData=str(tempData.value)
            if(tempData=='' or tempData is None):
                continue
            xlWorkBookData.append([index[0],rowIndex,colIndex,tempData])
            temptrans=translator.translate(tempData,dest='ja',src='en')
            temptrans=temptrans.text
            xlWorkSheetLst.append([index[0],rowIndex,colIndex,temptrans])
xlWorkBookMetaData=[]
print(xlWorkBookData)
print(xlWorkSheetLst)
for index in xlWorkSheetLst:
    xlWorkSheet=xlWorkbook[index[0]]
    transdata=xlWorkSheet.cell(row=index[1],column=index[2])
    transdata.value=index[3]
xlWorkbook.save("E:\\notes\\Python\\document translation tool\\testxl5.xlsx")